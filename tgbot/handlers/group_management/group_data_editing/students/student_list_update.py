import random
import aiohttp
import asyncio
import io
from aiogram.dispatcher import FSMContext
from aiogram.types import Message
from openpyxl import load_workbook

import tgbot.keyboards.reply as rkb
from tgbot.misc.login_generator import generate_login
from tgbot.handlers.group_management.group_data_editing.group_edit_form import send_group_edit_form
from tgbot.models.database_instance import db

group_name_request = "<b>Введите название группы, которую хотите обновить</b>"
failure_message = "Данные не удалось обработать, попробуйте позже..."

# Показывать актуальный список студентов после обновления

# Возможно стоит вынести эту функцию в отдельный файл



async def check_year_of_enrollment(message: Message, state: FSMContext):
    state_data = await state.get_data()
    group_name = state_data["group_name"]
    if group_name[:2] != "19" and group_name[-2:] != "19":
        await open_performance_list(message, state)
    else:
        del_msg = await message.answer("Обновление списка студентов групп 19 года не доступно",
                                       reply_markup=rkb.manager_keyboard)
        await state.update_data(del_msg=del_msg)
        await send_group_edit_form(message, state)


async def open_performance_list(message: Message, state: FSMContext):
    del_msg = await message.answer("Происходит обновление...\n\n"
                                   "Это может занять несколько секунд")
    await state.update_data(del_msg=del_msg)
    data = await state.get_data()
    group_name = data.get("group_name")
    performance_list_url = await db.get_performance_list_by_group_name(group_name)
    file_id = performance_list_url.split('/')[-2]
    file_url = f'https://drive.google.com/u/0/uc?id={file_id}&export=download'
    failure_flag = 0
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(file_url) as response:
                response.raise_for_status()
                workbook = load_workbook(filename=io.BytesIO(await response.read()))
    except Exception as e:
        await message.answer(failure_message)
        print(f"Unexpected error: {e}")
        failure_flag = 1
    if not failure_flag:
        await find_rows_and_columns(message, state, workbook)


async def find_rows_and_columns(message: Message, state: FSMContext, workbook):
    sheet = workbook.active
    target_row = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True), start=1):
        if row[0] is not None:
            target_row = row_number
            break

    column_mapping = {
        "№ п/п" : 0,
        "статус" : 1,
        "Фамилия": 5,
        "Имя": 6,
        "Отчество": 7
    }
    for cell in sheet[target_row]:
        cell_value = cell.value
        if cell_value in column_mapping:
            column_mapping[cell_value] = cell.column - 1

    for cell in sheet[target_row]:
        if cell.value == '№ п/п':
            column_letter = cell.column_letter
            column_values = sheet[column_letter]
            for column_cell in column_values:
                if column_cell.value == 1:
                    target_row = column_cell.row
                    break
            break
    await state.update_data(target_row=target_row,
                            num_col=column_mapping.get("№ п/п"),
                            status_col=column_mapping.get("статус"),
                            last_name_col=column_mapping.get("Фамилия"),
                            first_name_col=column_mapping.get("Имя"),
                            middle_name_col=column_mapping.get("Отчество"))


    await get_students_from_performance_list(message, state, sheet)

    
async def get_students_from_performance_list(message: Message, state: FSMContext, sheet):
    state_data = await state.get_data()
    num_col = state_data["num_col"]
    status_col = state_data["status_col"]
    last_name_col = state_data["last_name_col"]
    first_name_col = state_data["first_name_col"]
    middle_name_col = state_data["middle_name_col"]
    target_row = state_data["target_row"]

    learning_students = []
    async def process_row(row):
        num = row[num_col] if row[num_col] is not None else None
        status = row[status_col].strip() if row[status_col] is not None else None
        if status == "учится" and num is not None:
            last_name = row[last_name_col].strip() if row[last_name_col] is not None else None
            first_name = row[first_name_col].strip() if row[first_name_col] is not None else None
            middle_name = row[middle_name_col].strip() if row[middle_name_col] is not None else None

            learning_students.append([last_name, first_name, middle_name])
    
    rows = sheet.iter_rows(min_row=target_row, values_only=True)
    await asyncio.gather(*[process_row(row) for row in rows])
    await update_students_in_database(message, state, learning_students)



async def update_students_in_database(message: Message, state: FSMContext, learning_students):
    state_data = await state.get_data()
    group_name = state_data["group_name"]
    current_students = await db.get_students_by_group_name(group_name)
    current_student_list =[]
    for student in current_students:
        current_student_list.append([[student[0], student[1], student[2]], student[3]])

    year = group_name[-2:]
    added_students_data = []
    deleted_students = []

    for student in current_student_list:
        if student[0] not in learning_students:
            await db.delete_user_by_login(student[1])
            if student[0][2] is not None:
                deleted_students.append(f"◦ {student[0][0]} {student[0][1]} {student[0][2]};\n")
            else:
                deleted_students.append(f"◦ {student[0][0]} {student[0][1]};\n")

    for student in learning_students:
        is_added = 0
        for student_data in current_student_list:
            if student == student_data[0]:
                is_added = 1
                break
        if is_added == 0:
            login = await generate_login(student, year)
            password = random.randint(100000, 999999)
            user_type = "student"
            await db.add_user(student, login, password, user_type, group_name)
            if student[2] is not None:
                added_students_data.append(f"◦ {student[0]} {student[1]} {student[2]} : {login} : {password};\n")
            else:
                added_students_data.append(f"◦ {student[0]} {student[1]} : {login} : {password};\n")
    # Писать, если ничего не изменилось
    added_students_data = "".join(added_students_data)
    deleted_students = "".join(deleted_students)
    del_msg = state_data["del_msg"]
    await del_msg.delete()
    await message.answer(f"Группа {group_name} обновлена\n"
                         f"*Новые студенты:* \n{added_students_data}\n"
                         f"*Удаленные студенты:*\n{deleted_students}",
                         parse_mode="MARKDOWN",
                         reply_markup=rkb.manager_keyboard)
    await state.finish()
    
    
