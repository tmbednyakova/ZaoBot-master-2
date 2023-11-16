from aiogram import types
from aiogram.dispatcher import FSMContext
import openpyxl
import io
from urllib.request import urlopen
import random

import tgbot.keyboards.reply as rkb
from tgbot.filters.user_type import UserTypeFilter
from tgbot.misc.states import GroupAddStates
from tgbot.models.database_instance import db
from tgbot.misc.login_generator import generate_login


group_name_request = "<b>Введите название группы</b>"


async def start_adding_group(message: types.Message):
    bot = message.bot
    chat_id = message.from_user.id
    await bot.send_message(text=group_name_request, chat_id=chat_id, reply_markup=rkb.group_name_cancel_keyboard)
    await GroupAddStates.get_group_name_state.set()


async def check_group_name(message: types.Message, state: FSMContext):
    bot = message.bot
    chat_id = message.from_user.id
    new_group_name = message.text
    group_name_list = await db.get_group_names()
    is_duplicate = 0
    for group_name in group_name_list:
        if group_name[0] == new_group_name:
            await bot.send_message(
                text="<b>Группа с таким названием уже существует!</b>\n\n<b>Введите название группы</b>",
                chat_id=chat_id)
            is_duplicate = 1
            break
    if not is_duplicate:
        await state.update_data(new_group_name=new_group_name)
        await bot.send_message(text="<b>Название группы:</b> " + new_group_name +
                                    "\n\n<b>Введите адрес таблицы успеваемости, которую хотите добавить</b>",
                               chat_id=chat_id)
        await GroupAddStates.get_group_address_state.set()


async def check_link_address(message: types.Message, state: FSMContext):
    bot = message.bot
    data = await state.get_data()
    chat_id = message.from_user.id
    new_link = message.text
    link_list = await db.get_performance_urls()
    is_duplicate = 0
    for link in link_list:
        if link[0] == new_link:
            link_name = await db.get_group_name_by_url(new_link)
            new_group_name = data.get("new_group_name")
            await bot.send_message(
                text=f"<b>Такая ссылка уже привязана к группе</b> " + link_name + "\n\n<b>Название группы:</b> " +
                     new_group_name + "\n<b>Введите адрес таблицы успеваемости, которую хотите добавить</b>",
                chat_id=chat_id)
            await GroupAddStates.get_group_address_state.set()
            is_duplicate = 1
            break
    if not is_duplicate:
        await state.update_data(new_link=new_link)
        await add_group_to_database(message, state)


async def add_group_to_database(message: types.Message, state: FSMContext):
    bot = message.bot
    chat_id = message.from_user.id
    data = await state.get_data()
    new_link = data.get("new_link")
    new_group_name = data.get("new_group_name")
    await db.add_group(new_group_name, new_link)
    await bot.send_message(text="<b>Ссылка успешно добавлена</b>", chat_id=chat_id, reply_markup=rkb.empty_keyboard)
    await bot.send_message(text="<b>Происходит добавление студентов в базу данных...</b>",
                           chat_id=chat_id, reply_markup=rkb.manager_keyboard)
    await add_students_to_database(message, state)
    await state.finish()


async def add_students_to_database(message: types.Message, state: FSMContext):
    bot = message.bot
    data = await state.get_data()
    new_group_name = data.get("new_group_name")
    new_link = data.get("new_link")
    file_id = new_link.split('/')[-2]
    file_url = f'https://drive.google.com/u/0/uc?id={file_id}&export=download'
    wb = openpyxl.load_workbook(filename=io.BytesIO(urlopen(file_url).read()))

    students_data = ""
    if new_group_name[:2] == "19" or new_group_name[-2:] == "19":
        for sheet in wb.worksheets:
            if sheet.title.strip().endswith(".") or sheet.title.strip().endswith(")"):
                student = sheet.cell(row=1, column=2).value
                if student is None:
                    student = sheet.cell(row=1, column=3).value
                    if student is None:
                        print("Увы")
                        continue
                fio = student.split(" ")
                fio[0] = fio[0].capitalize()
                if len(fio) == 4:
                    fio = [fio[0] + " " + fio[1][0] + fio[1][1:].capitalize(), fio[2], fio[3]]
                year = new_group_name[-2:]
                if len(fio) < 3:
                    fio.append(None)
                login = await generate_login(fio, year)
                password = random.randint(100000, 999999)
                await db.add_user(list(fio), login, password, "student", new_group_name)
                if fio[2] is not None:
                    students_data += fio[0] + " " + fio[1] + " " + fio[2] + " " + login + " " + str(password) + "\n"
                else:
                    students_data += fio[0] + " " + fio[1] + " " + login + " " + str(password) + "\n"
    else:
        sheet = wb.active
        target_row = 0
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None:
                    target_row = cell.row
                    break
            else:
                continue
            break

        # Название колонки иногда отличается
        for cell in sheet[target_row]:
            if cell.value == '№':
                column_letter = cell.column_letter
                column_values = sheet[column_letter]
                for column_cell in column_values:
                    if column_cell.value == 1:
                        target_row = column_cell.row
                break

        unique_fio = set()
        for row in sheet.iter_rows(min_row=target_row, values_only=True):
            if row[0] is not None and row[1] == "учится":
                if row[5] is not None:
                    last_name = row[5].strip()
                else:
                    last_name = row[5]
                if row[6] is not None:
                    first_name = row[6].strip()
                else:
                    first_name = row[6]
                if row[7] is not None:
                    middle_name = row[7].strip()
                else:
                    middle_name = row[7]
                if all((last_name, first_name)):
                    fio = (last_name, first_name, middle_name)
                    if fio not in unique_fio:
                        unique_fio.add(fio)

        # Вынести в отдельную функцию
        # Переделать номер для экономики
        year = new_group_name[-2:]
        for fio in unique_fio:
            user = []
            for name in fio:
                user.append(name)
            if len(fio) < 3:
                user.append(None)
            login = await generate_login(user, year)
            password = random.randint(100000, 999999)
            await db.add_user(list(fio), login, password, "student", new_group_name)
            if fio[2] is not None:
                students_data += fio[0] + " " + fio[1] + " " + fio[2] + " " + login + " " + str(password) + "\n"
            else:
                students_data += fio[0] + " " + fio[1] + " " + login + " " + str(password) + "\n"

    await bot.send_message(chat_id=message.from_user.id,
                           text="Студенты успешно добавлены!\n"
                                f"<b>Данные студентов:</b> \n{students_data}\n\n",
                           reply_markup=rkb.manager_keyboard)
    await state.finish()



def register_group_add(dp):
    dp.register_message_handler(start_adding_group, UserTypeFilter("manager"), content_types=['text'],
                                text='Добавить группу')
    dp.register_message_handler(check_group_name, state=GroupAddStates.get_group_name_state)
    dp.register_message_handler(check_link_address, state=GroupAddStates.get_group_address_state)
