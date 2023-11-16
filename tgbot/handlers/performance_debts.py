"""Содержит функционал для просмотра успеваемости и долгов для студентов."""
import io
import aiohttp
import asyncio
from openpyxl import load_workbook
from openpyxl.utils import cell
from aiogram.types import Message
from tgbot.filters.user_type import UserTypeFilter
from tgbot.models.database_instance import db


failure_message = "Данные не удалось обработать, попробуйте позже..."

# Удалять это!
async def send_data_processing_notification(message: Message):
    """Отправляет оповещение о начале обработки данных."""
    await message.delete()
    await message.answer("Данные обрабатываются...\n"
                         "Это может занять несколько секунд")
    await prepare_excel_file(message)
    

# Возможно стоит подключаться к таблице при запуске бота и не отключаться от нее
# Создать словарь с группами и подключениями и проверить (узнать сколько места в памяти будут занимать книги)
async def prepare_excel_file(message: Message):
    """Подготавливает эксель таблицу группы для дальнейшей работы."""
    group_name = await db.get_user_group_name(message.from_user.id)
    file_url = await db.get_performance_list_by_group_name(group_name)
    file_id = file_url.split('/')[-2]
    file_url = f'https://drive.google.com/u/0/uc?id={file_id}&export=download'
    failure_flag = 0
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(file_url) as response:
                response.raise_for_status()
                workbook = load_workbook(filename=io.BytesIO(await response.read()))
    except Exception as e:
        await message.answer(failure_message)
        print(f"Unexpected error: {e}")
        failure_flag = 1
    if not failure_flag:
        await check_year_of_enrollment(message, group_name, workbook)


async def check_year_of_enrollment(message: Message, group_name, workbook):
    """Проверяет является ли группа 19 года поступления."""
    if group_name[:2] != "19" and group_name[-2:] != "19":
        await find_rows_and_columns(message, workbook)
    else:
        await go_through_group_excel_sheets_19(message, workbook)


async def find_rows_and_columns(message: Message, workbook):
    """Определяет столбцы и строки с искомыми значениями."""
    sheet = workbook.active
    target_row = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True), start=1):
        if row[0] is not None:
            target_row = row_number
            break
    
    # Возможно следует занулить эти значения, чтоб компактнее объявлять словарь
    column_mapping = {
        "Фамилия": 5,
        "Имя": 6,
        "Отчество": 7,
        "Наименование дисциплин": 8,
        "Отметка о сдаче": 11,
        "курс": 3
    }

    for cell in sheet[target_row]:
        cell_value = cell.value
        if cell_value in column_mapping:
            column_mapping[cell_value] = cell.column - 1
            
    for cell in sheet[target_row]:
        if cell.value == '№ п/п':
            column_letter = cell.column_letter
            column_values = sheet[column_letter]
            for column_cell in column_values:
                if column_cell.value == 1:
                    target_row = column_cell.row
                    break
            break
    
    last_name_col = column_mapping.get("Фамилия")
    first_name_col = column_mapping.get("Имя")
    middle_name_col = column_mapping.get("Отчество")
    mark_col = column_mapping.get("Отметка о сдаче")
    subject_name_col = column_mapping.get("Наименование дисциплин")
    semestra_col = column_mapping.get("курс")

    if message.text == "Успеваемость":
        await create_performance_message(message, sheet, target_row, last_name_col, first_name_col, middle_name_col, 
                                        mark_col, subject_name_col)
    elif message.text == "Долги":
        await create_debt_message(message, sheet, target_row, last_name_col, first_name_col, middle_name_col, 
                                  mark_col, subject_name_col, semestra_col)

# Проверять только определенное количество последовательных значений, после того как нашел первую подходящую строку
async def create_performance_message(message: Message, sheet, target_row, 
                                     last_name_col, first_name_col, middle_name_col, 
                                     mark_col, subject_name_col):
    """Собирает данные по успеваемости студента, вызвавшего функцию и создает сообщение."""
    user_last_name = await db.get_user_ln(message.from_user.id)
    user_first_name = await db.get_user_fn(message.from_user.id)
    user_middle_name = await db.get_user_mn(message.from_user.id)

    user_last_name = user_last_name[0]
    user_first_name = user_first_name[0]
    if user_middle_name is not None:
        user_middle_name = user_middle_name[0]

    performance_rows = []
    
    async def process_row(row):
        last_name = row[last_name_col].strip() if row[last_name_col] is not None and isinstance(row[last_name_col], str) else None
        first_name = row[first_name_col].strip() if row[first_name_col] is not None and isinstance(row[first_name_col], str) else None
        middle_name = row[middle_name_col].strip() if row[middle_name_col] is not None and isinstance(row[middle_name_col], str) else None

        if last_name == user_last_name and first_name == user_first_name and middle_name == user_middle_name:
            mark = row[mark_col].strip() if row[mark_col] is not None and isinstance(row[mark_col], str) else None
            subject_name = row[subject_name_col].strip() if row[subject_name_col] is not None and isinstance(row[subject_name_col], str) else None

            if mark != '' and mark != 'долг' and mark is not None:
                performance_rows.append(f"◦ {subject_name} - <b>{mark}</b>")

    
    rows = sheet.iter_rows(min_row=target_row, values_only=True)
    await asyncio.gather(*[process_row(row) for row in rows])
    
    performance_message = "<b>Ваша успеваемость:</b>\n" + "\n".join(performance_rows)
    await send_performance_message(message, performance_message)


async def create_debt_message(message: Message, sheet, target_row,
                              last_name_col, first_name_col, middle_name_col,
                              mark_col, subject_name_col, semestra_col):
    """Собирает данные по долгам студента, вызвавшего функцию и создает сообщение."""
    # Доделать!
    user_last_name = await db.get_user_ln(message.from_user.id)
    user_first_name = await db.get_user_fn(message.from_user.id)
    user_middle_name = await db.get_user_mn(message.from_user.id)
    
    user_last_name = user_last_name[0]
    user_first_name = user_first_name[0]
    if user_middle_name is not None:
        user_middle_name = user_middle_name[0]

    debt_rows = []

    async def process_row(row):
        last_name = row[last_name_col].strip() if row[last_name_col] is not None and isinstance(row[last_name_col], str) else None
        first_name = row[first_name_col].strip() if row[first_name_col] is not None and isinstance(row[first_name_col], str) else None
        middle_name = row[middle_name_col].strip() if row[middle_name_col] is not None and isinstance(row[middle_name_col], str) else None

        if last_name == user_last_name and first_name == user_first_name and middle_name == user_middle_name:
            mark = row[mark_col].strip() if row[mark_col] is not None and isinstance(row[mark_col], str) else None
            subject_name = row[subject_name_col].strip() if row[subject_name_col] is not None and isinstance(row[subject_name_col], str) else None
            semestra = row[semestra_col].strip() if row[semestra_col] is not None and isinstance(row[semestra_col], str) else None

            if mark == 'долг':
                debt_rows.append(f"◦ {subject_name} - <b>{semestra}</b>")


    rows = sheet.iter_rows(min_row=target_row, values_only=True)
    await asyncio.gather(*[process_row(row) for row in rows])
    
    debt_message = "<b>Ваши долги:</b>\n" + "\n".join(debt_rows)
    await send_debt_message(message, debt_message)
    

async def go_through_group_excel_sheets_19(message: Message, workbook):
    """Проходится по листам таблицы и вызывает проверку на совпадение имени студента с названием листа."""
    for sheet in workbook.worksheets:
        sheet_title = sheet.title.strip()
        if sheet_title.endswith(".") or sheet_title.endswith(")"):
            for col in [2, 3]:
                student_name = sheet[cell.get_column_letter(col) + '1'].value
                if student_name is not None:
                    await check_student_name_19(message, sheet, student_name)
                    break
            else:
                print("Имя студента не обнаружено")
                continue
            

async def check_student_name_19(message: Message, sheet, student_name):
    """Проверяет совпадает ли имя студента вызвавшего функцию с названием листа."""
    student_name_parts = student_name.split(" ")
    if len(student_name_parts[0].split("-")) == 1:
        student_name_parts[0] = student_name_parts[0].capitalize()
    else:
        arabian_last_name = student_name_parts[0].split("-")
        arabian_last_name = [part.capitalize() for part in arabian_last_name]
        student_name_parts[0] = "-".join(arabian_last_name)
    if len(student_name_parts) == 4:
        student_name_parts = [student_name_parts[0] + " " + student_name_parts[1][0] + student_name_parts[1][1:].capitalize(),
                                student_name_parts[2], student_name_parts[3]]
    user_name = await db.get_fio(message.from_user.id)
    student_name = " ".join(student_name_parts)
    if user_name == student_name:
        await find_rows_and_columns_19(message, sheet)


async def find_rows_and_columns_19(message: Message, sheet):
    """Находит столбцы и колонки с искомыми значениями."""
    known_cells = [(5, 3), (6, 2)] 
    for row, col in known_cells:
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value == "Наименование дисциплин":
            target_row = row + 1
            subject_name_col = col - 1
            mark_col = subject_name_col + 3
            if message.text == "Долги":
                await create_debt_message_19(message, sheet, target_row, subject_name_col, mark_col)
                break
            elif message.text == "Успеваемость":
                await create_performance_message_19(message, sheet, target_row, subject_name_col, mark_col)
                break


async def create_debt_message_19(message: Message, sheet, target_row, subject_name_col, mark_col):
    """Создает сообщение с долгами студента."""
    debt_message = "<b>Ваши долги:</b>\n"
    for row in sheet.iter_rows(min_row=target_row, values_only=True):
        if row[mark_col] is not None and row[mark_col] == 'долг':
            debt_message += f"◦ {row[subject_name_col].strip()}\n"
    await send_debt_message(message, debt_message)


async def create_performance_message_19(message: Message, sheet, target_row, subject_name_col, mark_col):
    """Создает сообщение с успеваемостью студента."""
    performance_message = "<b>Ваша успеваемость:</b>\n"
    for row in sheet.iter_rows(min_row=target_row, values_only=True):
        if row[mark_col] is not None and row[mark_col] != 'долг':
            performance_message += f"◦ {row[subject_name_col].strip()} - <b>{row[mark_col].strip()}</b>\n"
    await send_performance_message(message, performance_message)


async def send_debt_message(message: Message, debt_message):
    """Отправляет сообщение со списком долгов студенту."""
    if debt_message == "<b>Ваши долги:</b>\n":
        debt_message = "<b>У вас нет долгов</b>"
    await message.answer(debt_message)


# Если есть долги, то писать, что есть только долги
async def send_performance_message(message: Message, performance_message):
    """Отправляет сообщение с успеваемостью студенту."""
    if performance_message == "<b>Ваша успеваемость:</b>\n":
        performance_message = "<b>Ваша успеваемость еще не сформирована</b>"
    await message.answer(performance_message)



def register_performance_debts_check(dp):
    dp.register_message_handler(send_data_processing_notification, UserTypeFilter("student"),
                                content_types=['text'], text=['Успеваемость'])
    dp.register_message_handler(send_data_processing_notification, UserTypeFilter("student"),
                                content_types=['text'], text=['Долги'])